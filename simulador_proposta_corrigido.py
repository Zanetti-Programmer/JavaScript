import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Criando um Excel para comparação de propostas
wb = Workbook()
ws = wb.active
ws.title = "Simulador"

# Cabeçalhos
headers = ["Item", "Empresa Atual (R$)", "Nova Proposta (R$)", "Diferença (R$)"]
ws.append(headers)

# Dados base (empresa atual) - corrigidos
data = [
    ["Salário", 3500, "", "=C2-B2"],
    ["Vale Refeição (VR)", 550, "", "=C3-B3"],
    ["Vale Alimentação (VA)", 504, "", "=C4-B4"],
    ["Vale Transporte (VT)", 0, "", "=C5-B5"],
    ["TotalPass", 0, "", "=C6-B6"],  # Corrigido: valor numérico
    ["Plano de Saúde", 0, "", "=C7-B7"],  # Valor 0 para "sem coparticipação"
    ["Outros (PLR, bônus...)", 0, "", "=C8-B8"],
    ["", "", "", ""],  # Linha em branco para separar
    ["TOTAL Mensal", "=B2+B3+B4+B5+B6+B7+B8", "=C2+C3+C4+C5+C6+C7+C8", "=C10-B10"],
]

# Adicionando os dados
for i, row in enumerate(data, start=2):
    for j, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j)
        cell.value = value

# Formatação dos cabeçalhos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, 5):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Formatação da linha de total
total_font = Font(bold=True)
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for col in range(1, 5):
    cell = ws.cell(row=10, column=col)  # Linha do total
    cell.font = total_font
    cell.fill = total_fill

# Ajustar largura das colunas
column_widths = [25, 20, 20, 18]
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Formatação de números como moeda brasileira
from openpyxl.styles import NamedStyle

# Criar estilo de moeda
currency_style = NamedStyle(name="currency")
currency_style.number_format = 'R$ #,##0.00'

# Aplicar formatação de moeda nas colunas B, C e D (exceto cabeçalho e linha vazia)
for row in range(2, 11):  # linhas 2 a 10
    if row != 9:  # pular linha vazia
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].style = currency_style

# Adicionar bordas
from openpyxl.styles import Border, Side

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Aplicar bordas em toda a tabela
for row in range(1, 11):
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = thin_border

# Adicionar instruções
ws["A12"] = "INSTRUÇÕES:"
ws["A13"] = "1. Preencha os valores da 'Nova Proposta' na coluna C"
ws["A14"] = "2. As diferenças serão calculadas automaticamente"
ws["A15"] = "3. Valores positivos na coluna 'Diferença' indicam aumento"
ws["A16"] = "4. Valores negativos indicam redução"

# Formatação das instruções
instruction_font = Font(italic=True, size=10)
for row in range(12, 17):
    ws[f'A{row}'].font = instruction_font

# Salvar arquivo no diretório atual
try:
    wb.save("Simulador_Proposta.xlsx")
    print("✅ Arquivo 'Simulador_Proposta.xlsx' criado com sucesso!")
    print("📁 Localização: diretório atual")
except Exception as e:
    print(f"❌ Erro ao salvar arquivo: {e}")

# Mostrar resumo da planilha criada
print("\n📊 RESUMO DA PLANILHA:")
print("- Simulador de comparação de propostas de emprego")
print("- Campos: Salário, VR, VA, VT, TotalPass, Plano de Saúde, Outros")
print("- Cálculos automáticos de diferença e total")
print("- Formatação profissional com cores e bordas")
print("- Formato de moeda brasileira (R$)")
